from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
wb = Workbook()
ws = wb.active
def styling():
	count = 1
	fonts = Font(name="Cambria", size=14)	
	global_border = Border(top=Side(border_style="thin"), bottom=Side(border_style="thin"), right=Side(border_style="thin"), left=Side(border_style="thin"))
	center_alignment = Alignment(horizontal="center")
	Worksheet.set_printer_settings(ws, paper_size=5, orientation="landscape")
	ws.page_margins.left = 0.5
	ws.page_margins.top = 0.5
	ws["A1"].alignment = center_alignment
	while ws[get_column_letter(count) + "3"].value:
		ws[get_column_letter(count) + "3"].alignment = center_alignment
		count += 1
	ws.merge_cells("A1:" + get_column_letter(count-1) + "1")
	ws.merge_cells("A2:" + get_column_letter(count-1) + "2")
	for row in range(1, count+3):
		for col in range(1, count):
			ws[get_column_letter(col) + str(row)].border = global_border
			ws[get_column_letter(col) + str(row)].font = fonts
			if col == 1 or col == 3 and row == 1 or row == 2:
				ws.column_dimensions[get_column_letter(col)].width = 20
			else:
				ws.column_dimensions[get_column_letter(col)].width = 18
def main():
	section = ["Product Name", "Prev. Inv.", "Pull-Out Stocks", "Total Stocks", "SOH", "Stocks Sold", "Unit", "Amount"]
	ws["A1"] = "Name of Company"	
	ws["A2"] = "Supplier:                                                                                                                                                Date: __/__/____" 
	ws.append(section)
	styling()
	wb.save("C:\\Users\\JessG\\Desktop\\Mom's Work\\test.xlsx")
if __name__ == '__main__':
	main()